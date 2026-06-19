#!/usr/bin/env python3
# 生成数据收集进展 Excel 文档，供工作组展示
# 输出: plant_dust_v2/数据收集进展.xlsx
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart
from collections import Counter
import os, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

ROOT = os.path.dirname(os.path.abspath(__file__))
CSV = os.path.join(ROOT, 'data', 'dataset.csv')
OUT = os.path.join(ROOT, 'output', '数据收集进展.xlsx')
# DESKTOP_OUT = r'c:\Users\汪昱全\Desktop\数据收集进展.xlsx'  # 笔记本专用，本机注释

df = pd.read_csv(CSV, encoding='utf-8-sig')

# ============================================================
# 样式定义
# ============================================================
HEADER_FONT = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
SUBHEADER_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
SUBHEADER_FONT = Font(name='微软雅黑', bold=True, size=10)
BODY_FONT = Font(name='微软雅黑', size=10)
TITLE_FONT = Font(name='微软雅黑', bold=True, size=14, color='1F3864')
SECTION_FONT = Font(name='微软雅黑', bold=True, size=12, color='2F5496')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)

# 颜色方案
GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
RED_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
LIGHT_BLUE = PatternFill(start_color='DAEEF3', end_color='DAEEF3', fill_type='solid')

def style_header(ws, row, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

def style_body(ws, start_row, end_row, ncols):
    for row in range(start_row, end_row + 1):
        for col in range(1, ncols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = BODY_FONT
            cell.alignment = CENTER
            cell.border = THIN_BORDER

def auto_width(ws, min_width=8, max_width=50):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = min_width
        for cell in col:
            if cell.value:
                lines = str(cell.value).split('\n')
                for line in lines:
                    # 中文字符算2个宽度
                    length = sum(2 if ord(c) > 127 else 1 for c in line)
                    max_len = max(max_len, length)
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)

# ============================================================
# 创建工作簿
# ============================================================
wb = Workbook()

# ---- Sheet 1: 数据总览 ----
ws1 = wb.active
ws1.title = '数据总览'

# 大标题
ws1.merge_cells('A1:H1')
ws1.cell(row=1, column=1, value='植物叶片滞尘数据集 — 数据收集进展').font = TITLE_FONT
ws1.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 30

ws1.merge_cells('A2:H2')
ws1.cell(row=2, column=1, value=f'报告日期：2026-06-18 | 方法论：物种优先 × 跨功能区对比 | 目标区域：华中+华东').font = Font(name='微软雅黑', size=9, color='666666')
ws1.cell(row=2, column=1).alignment = CENTER

# 核心指标卡片
row = 4
ref_count = df['reference'].nunique()
metrics = [
    ('总记录数', str(len(df)), '条'),
    ('目标物种', '14', '种（5乔木+5灌木+4地被）'),
    ('覆盖城市', str(df['city'].nunique()), '个'),
    ('参考文献', str(ref_count), '篇'),
    ('含TSP数据', f'{df["tsp_g_m2"].notna().sum() if "tsp_g_m2" in df.columns else 0}', '条'),
    ('含PM10数据', f'{df["pm10_g_m2"].notna().sum() if "pm10_g_m2" in df.columns else 0}', '条'),
    ('含PM2.5数据', f'{df["pm2_5_g_m2"].notna().sum() if "pm2_5_g_m2" in df.columns else 0}', '条'),
    ('功能区类型', '6', '工业区/交通干道/公园清洁区/居住区/城市混合/文教区'),
]

for i, (label, value, unit) in enumerate(metrics):
    c = (i % 4) * 2 + 1
    r = row + (i // 4) * 2
    ws1.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c+1)
    cell = ws1.cell(row=r, column=c, value=label)
    cell.font = Font(name='微软雅黑', size=10, color='666666')
    cell.alignment = CENTER
    ws1.merge_cells(start_row=r+1, start_column=c, end_row=r+1, end_column=c+1)
    cell = ws1.cell(row=r+1, column=c, value=f'{value} {unit}')
    cell.font = Font(name='微软雅黑', bold=True, size=16, color='2F5496')
    cell.alignment = CENTER
    for rr in [r, r+1]:
        for cc in [c, c+1]:
            ws1.cell(row=rr, column=cc).fill = LIGHT_BLUE
            ws1.cell(row=rr, column=cc).border = THIN_BORDER

# ---- 物种数据量统计 ----
row = 10
ws1.merge_cells(f'A{row}:H{row}')
ws1.cell(row=row, column=1, value='■ 目标物种数据量').font = SECTION_FONT

row = 11
species_headers = ['层级', '物种', '学名', '记录数', 'TSP', 'PM10', 'PM2.5', '覆盖城市']
for i, h in enumerate(species_headers, 1):
    ws1.cell(row=row, column=i, value=h)
style_header(ws1, row, len(species_headers))

species_data = [
    ('乔木', '香樟', 'Cinnamomum camphora'),
    ('乔木', '桂花', 'Osmanthus fragrans'),
    ('乔木', '二球悬铃木', 'Platanus acerifolia'),
    ('乔木', '广玉兰', 'Magnolia grandiflora'),
    ('乔木', '女贞', 'Ligustrum lucidum'),
    ('灌木', '海桐', 'Pittosporum tobira'),
    ('灌木', '红叶石楠', 'Photinia × fraseri'),
    ('灌木', '红花檵木', 'Loropetalum chinense var. rubrum'),
    ('灌木', '杜鹃', 'Rhododendron simsii'),
    ('灌木', '石楠', 'Photinia serrulata'),
    ('地被', '八角金盘', 'Fatsia japonica'),
    ('地被', '洒金桃叶珊瑚', 'Aucuba japonica var. variegata'),
    ('地被', '麦冬/沿阶草', 'Ophiopogon japonicus/bodinieri'),
]

row = 12
for layer, cn, latin in species_data:
    mask = df['plant_species'].apply(lambda x: x in cn.replace('/','/').split('/') if '/' in cn else x == cn)
    # 处理麦冬/沿阶草
    if '/' in cn:
        parts = cn.split('/')
        mask = df['plant_species'].isin(parts)
    else:
        mask = df['plant_species'] == cn

    sub = df[mask]
    n = len(sub)
    tsp_n = sub['tsp_g_m2'].notna().sum()
    pm10_n = sub['pm10_g_m2'].notna().sum()
    pm25_n = sub['pm2_5_g_m2'].notna().sum()
    cities = ', '.join(sorted(sub['city'].dropna().unique()))

    vals = [layer, cn, latin, n, tsp_n, pm10_n, pm25_n, cities]
    for i, v in enumerate(vals, 1):
        ws1.cell(row=row, column=i, value=v)
    row += 1

style_body(ws1, 12, row - 1, len(species_headers))

# 合计行
ws1.cell(row=row, column=1, value='合计').font = Font(name='微软雅黑', bold=True, size=10)
ws1.cell(row=row, column=3, value='14种')
ws1.cell(row=row, column=4, value=len(df))
ws1.cell(row=row, column=5, value=df['tsp_g_m2'].notna().sum())
ws1.cell(row=row, column=6, value=df['pm10_g_m2'].notna().sum())
ws1.cell(row=row, column=7, value=df['pm2_5_g_m2'].notna().sum())
for col in range(1, 9):
    ws1.cell(row=row, column=col).font = Font(name='微软雅黑', bold=True, size=10)
    ws1.cell(row=row, column=col).fill = SUBHEADER_FILL
    ws1.cell(row=row, column=col).border = THIN_BORDER
    ws1.cell(row=row, column=col).alignment = CENTER

# ---- 城市数据量统计 ----
row += 2
ws1.merge_cells(f'A{row}:H{row}')
ws1.cell(row=row, column=1, value='■ 城市分布').font = SECTION_FONT

row += 1
city_headers = ['城市', '省份', '区域分类', '记录数', '乔木', '灌木', '地被', '覆盖物种数']
for i, h in enumerate(city_headers, 1):
    ws1.cell(row=row, column=i, value=h)
style_header(ws1, row, len(city_headers))

city_meta = {
    '武汉': ('湖北', '华中'), '杭州': ('浙江', '华东'), '长沙': ('湖南', '华中'),
    '郑州': ('河南', '华中'), '南昌': ('江西', '华中'), '扬州': ('江苏', '华东'),
    '南京': ('江苏', '华东'), '上海': ('上海', '华东'), '合肥': ('安徽', '华中'),
    '洛阳': ('河南', '华中'), '福州': ('福建', '华东'), '深圳': ('广东', '华东'),
}

row += 1
for city, (prov, region) in sorted(city_meta.items(), key=lambda x: -len(df[df['city']==x[0]])):
    sub = df[df['city'] == city]
    n = len(sub)
    tree = len(sub[sub['layer'] == '乔木'])
    shrub = len(sub[sub['layer'] == '灌木'])
    herb = len(sub[sub['layer'] == '地被'])
    spp = sub['plant_species'].nunique()

    vals = [city, prov, region, n, tree, shrub, herb, spp]
    for i, v in enumerate(vals, 1):
        ws1.cell(row=row, column=i, value=v)

    # 高亮华中城市
    if region == '华中':
        for col in range(1, len(city_headers) + 1):
            ws1.cell(row=row, column=col).fill = GREEN_FILL

    row += 1

style_body(ws1, row - 12, row - 1, len(city_headers))

# ---- 数据完整性指标 ----
row += 1
ws1.merge_cells(f'A{row}:H{row}')
ws1.cell(row=row, column=1, value='■ 数据完整性指标').font = SECTION_FONT

row += 1
integrity_headers = ['指标', '当前值', '说明']
for i, h in enumerate(integrity_headers, 1):
    ws1.cell(row=row, column=i, value=h)
style_header(ws1, row, len(integrity_headers))

integrity_data = [
    ('TSP覆盖率', f'{df["tsp_g_m2"].notna().sum()}/{len(df)} ({df["tsp_g_m2"].notna().sum()/len(df)*100:.0f}%)',
     '单位叶面积总悬浮颗粒物滞尘量（核心指标）'),
    ('PM10覆盖率', f'{df["pm10_g_m2"].notna().sum()}/{len(df)} ({df["pm10_g_m2"].notna().sum()/len(df)*100:.0f}%)',
     '可吸入颗粒物（空气动力学直径<10μm）'),
    ('PM2.5覆盖率', f'{df["pm2_5_g_m2"].notna().sum()}/{len(df)} ({df["pm2_5_g_m2"].notna().sum()/len(df)*100:.0f}%)',
     '细颗粒物（空气动力学直径<2.5μm）'),
    ('乔木:灌木:地被', f'{len(df[df["layer"]=="乔木"])}:{len(df[df["layer"]=="灌木"])}:{len(df[df["layer"]=="地被"])}',
     f'地被层{len(df[df["layer"]=="地被"])}条({len(df[df["layer"]=="地被"])/len(df)*100:.0f}%)，仍需重点补充'),
    ('华中:华东比例', f'{len(df[df["city"].isin(["武汉","长沙","郑州","南昌","洛阳","合肥"])])}:{len(df[df["city"].isin(["杭州","扬州","南京","上海","福州","深圳"])])}',
     '华中6城与华东6城数据分布（合肥计入华中）'),
]

row += 1
for metric, value, note in integrity_data:
    ws1.cell(row=row, column=1, value=metric)
    ws1.cell(row=row, column=2, value=value)
    ws1.cell(row=row, column=3, value=note)
    ws1.cell(row=row, column=3).alignment = LEFT
    row += 1

style_body(ws1, row - len(integrity_data), row - 1, len(integrity_headers))

auto_width(ws1)

# ---- Sheet 2: 物种×城市覆盖矩阵 ----
ws2 = wb.create_sheet('物种×城市覆盖矩阵')

TARGET_SPECIES = [
    ('香樟', '乔木'), ('桂花', '乔木'), ('二球悬铃木', '乔木'),
    ('广玉兰', '乔木'), ('女贞', '乔木'),
    ('海桐', '灌木'), ('红叶石楠', '灌木'), ('红花檵木', '灌木'), ('杜鹃', '灌木'), ('石楠', '灌木'),
    ('八角金盘', '地被'), ('洒金桃叶珊瑚', '地被'),
    ('麦冬', '地被'), ('沿阶草', '地被'),
]

CITIES_SORTED = ['武汉', '杭州', '长沙', '郑州', '南昌', '扬州', '南京', '上海', '合肥', '洛阳', '福州', '深圳']

ws2.merge_cells('A1:O1')
ws2.cell(row=1, column=1, value='物种 × 城市 数据覆盖矩阵').font = TITLE_FONT
ws2.cell(row=1, column=1).alignment = CENTER
ws2.row_dimensions[1].height = 30

ws2.merge_cells('A2:O2')
ws2.cell(row=2, column=1, value='单元格数字 = 该物种在该城市的记录条数 | 颜色深浅反映数据量（深绿=丰富，浅绿=有数据，灰色=无数据）').font = Font(name='微软雅黑', size=9, color='666666')
ws2.cell(row=2, column=1).alignment = CENTER

# 表头
row = 4
ws2.cell(row=row, column=1, value='物种')
ws2.cell(row=row, column=2, value='层级')
for i, city in enumerate(CITIES_SORTED):
    ws2.cell(row=row, column=3+i, value=city)
ws2.cell(row=row, column=3+len(CITIES_SORTED), value='合计')
style_header(ws2, row, 3 + len(CITIES_SORTED))

# 数据填充
DARK_GREEN = PatternFill(start_color='1B7837', end_color='1B7837', fill_type='solid')
MID_GREEN = PatternFill(start_color='5AAE61', end_color='5AAE61', fill_type='solid')
LIGHT_GREEN_FILL = PatternFill(start_color='A6DBA0', end_color='A6DBA0', fill_type='solid')
PALE_GREEN = PatternFill(start_color='D9F0D3', end_color='D9F0D3', fill_type='solid')
GREY_FILL = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
WHITE_FONT = Font(name='微软雅黑', bold=True, size=10, color='FFFFFF')

row = 5
for sp, layer in TARGET_SPECIES:
    ws2.cell(row=row, column=1, value=sp)
    ws2.cell(row=row, column=2, value=layer)

    # 合并麦冬/沿阶草的count
    if sp in ('麦冬', '沿阶草'):
        sub = df[df['plant_species'].isin(['麦冬', '沿阶草'])]
    else:
        sub = df[df['plant_species'] == sp]

    total = 0
    for j, city in enumerate(CITIES_SORTED):
        n = len(sub[sub['city'] == city])
        total += n
        cell = ws2.cell(row=row, column=3+j, value=n if n > 0 else '—')

        if n >= 10:
            cell.fill = DARK_GREEN
            cell.font = WHITE_FONT
        elif n >= 5:
            cell.fill = MID_GREEN
            cell.font = WHITE_FONT
        elif n >= 2:
            cell.fill = LIGHT_GREEN_FILL
        elif n >= 1:
            cell.fill = PALE_GREEN
        else:
            cell.fill = GREY_FILL

    ws2.cell(row=row, column=3+len(CITIES_SORTED), value=total)
    ws2.cell(row=row, column=3+len(CITIES_SORTED)).font = Font(name='微软雅黑', bold=True, size=10)

    row += 1

style_body(ws2, 5, row - 1, 3 + len(CITIES_SORTED))

# 合计行
ws2.cell(row=row, column=1, value='合计').font = Font(name='微软雅黑', bold=True, size=10)
for j, city in enumerate(CITIES_SORTED):
    total = len(df[df['city'] == city])
    cell = ws2.cell(row=row, column=3+j, value=total)
    cell.font = Font(name='微软雅黑', bold=True, size=10)
ws2.cell(row=row, column=3+len(CITIES_SORTED), value=len(df))
ws2.cell(row=row, column=3+len(CITIES_SORTED)).font = Font(name='微软雅黑', bold=True, size=10)
for col in range(1, 4 + len(CITIES_SORTED)):
    ws2.cell(row=row, column=col).fill = SUBHEADER_FILL
    ws2.cell(row=row, column=col).border = THIN_BORDER
    ws2.cell(row=row, column=col).alignment = CENTER

# 图例
row += 2
ws2.cell(row=row, column=1, value='图例：').font = Font(name='微软雅黑', bold=True, size=9)
legends = [
    (DARK_GREEN, '≥10条（数据丰富）'),
    (MID_GREEN, '5-9条（数据充足）'),
    (LIGHT_GREEN_FILL, '2-4条（数据稀疏）'),
    (PALE_GREEN, '1条（数据极少）'),
    (GREY_FILL, '无数据'),
]
for i, (fill, label) in enumerate(legends):
    c = 2 + i * 2
    ws2.cell(row=row, column=c).fill = fill
    ws2.cell(row=row, column=c).border = THIN_BORDER
    ws2.cell(row=row, column=c+1, value=label).font = Font(name='微软雅黑', size=9)

auto_width(ws2)

# ---- Sheet 3: 物种×功能区覆盖矩阵 ----
ws3 = wb.create_sheet('物种×功能区覆盖矩阵')

ZONES = ['工业区', '交通干道', '公园清洁区', '居住区', '城市混合', '文教区']
# 标准化映射（处理名称不一致）
def normalize_zone(z):
    z = str(z).strip()
    if '工业' in z: return '工业区'
    if '交通' in z or '道路' in z or '街道' in z: return '交通干道'
    if '公园' in z or '清洁' in z: return '公园清洁区'
    if '居住' in z: return '居住区'
    if '文教' in z or '校园' in z or '大学' in z: return '文教区'
    if '城市混合' in z or '混合' in z: return '城市混合'
    if '新兴' in z: return '城市混合'
    return z

df['zone_normalized'] = df['functional_zone'].apply(normalize_zone)

ws3.merge_cells('A1:I1')
ws3.cell(row=1, column=1, value='物种 × 功能区 数据覆盖矩阵').font = TITLE_FONT
ws3.cell(row=1, column=1).alignment = CENTER
ws3.row_dimensions[1].height = 30

ws3.merge_cells('A2:I2')
ws3.cell(row=2, column=1, value='单元格数字 = 该物种在该功能区的记录条数 | 跨功能区对比是核心方法论').font = Font(name='微软雅黑', size=9, color='666666')
ws3.cell(row=2, column=1).alignment = CENTER

row = 4
ws3.cell(row=row, column=1, value='物种')
ws3.cell(row=row, column=2, value='层级')
for i, z in enumerate(ZONES):
    ws3.cell(row=row, column=3+i, value=z)
ws3.cell(row=row, column=3+len(ZONES), value='合计')
style_header(ws3, row, 3 + len(ZONES))

row = 5
for sp, layer in TARGET_SPECIES:
    ws3.cell(row=row, column=1, value=sp)
    ws3.cell(row=row, column=2, value=layer)

    if sp in ('麦冬', '沿阶草'):
        sub = df[df['plant_species'].isin(['麦冬', '沿阶草'])]
    else:
        sub = df[df['plant_species'] == sp]

    total = 0
    for j, zone in enumerate(ZONES):
        n = len(sub[sub['zone_normalized'] == zone])
        total += n
        cell = ws3.cell(row=row, column=3+j, value=n if n > 0 else '—')
        if n >= 5:
            cell.fill = MID_GREEN
            cell.font = WHITE_FONT
        elif n >= 2:
            cell.fill = LIGHT_GREEN_FILL
        elif n >= 1:
            cell.fill = PALE_GREEN
        else:
            cell.fill = GREY_FILL

    ws3.cell(row=row, column=3+len(ZONES), value=total)
    ws3.cell(row=row, column=3+len(ZONES)).font = Font(name='微软雅黑', bold=True, size=10)
    row += 1

style_body(ws3, 5, row - 1, 3 + len(ZONES))

# 合计行
ws3.cell(row=row, column=1, value='合计').font = Font(name='微软雅黑', bold=True, size=10)
for j, zone in enumerate(ZONES):
    total = len(df[df['zone_normalized'] == zone])
    cell = ws3.cell(row=row, column=3+j, value=total)
    cell.font = Font(name='微软雅黑', bold=True, size=10)
ws3.cell(row=row, column=3+len(ZONES), value=len(df))
ws3.cell(row=row, column=3+len(ZONES)).font = Font(name='微软雅黑', bold=True, size=10)
for col in range(1, 4 + len(ZONES)):
    ws3.cell(row=row, column=col).fill = SUBHEADER_FILL
    ws3.cell(row=row, column=col).border = THIN_BORDER
    ws3.cell(row=row, column=col).alignment = CENTER

auto_width(ws3)

# ---- Sheet 4: 参考文献清单 ----
ws4 = wb.create_sheet('参考文献清单')

# 从数据集提取唯一参考文献
refs = df[['reference', 'doi']].drop_duplicates(subset='reference')
refs = refs[refs['reference'].notna() & (refs['reference'] != '')]

ws4.merge_cells('A1:E1')
ws4.cell(row=1, column=1, value=f'参考文献清单（共 {len(refs)} 篇）').font = TITLE_FONT
ws4.cell(row=1, column=1).alignment = CENTER
ws4.row_dimensions[1].height = 30

row = 3
ws4_headers = ['序号', '文献引用', 'DOI', '贡献记录数', '覆盖物种']
for i, h in enumerate(ws4_headers, 1):
    ws4.cell(row=row, column=i, value=h)
style_header(ws4, row, len(ws4_headers))

row = 4
for idx, (_, r) in enumerate(refs.iterrows(), 1):
    ref = r['reference']
    doi = r['doi'] if pd.notna(r['doi']) else ''
    sub = df[df['reference'] == ref]
    n = len(sub)
    spp = ', '.join(sorted(sub['plant_species'].unique()))

    ws4.cell(row=row, column=1, value=idx)
    ws4.cell(row=row, column=2, value=str(ref)[:120])
    ws4.cell(row=row, column=2).alignment = LEFT
    ws4.cell(row=row, column=3, value=doi)
    ws4.cell(row=row, column=4, value=n)
    ws4.cell(row=row, column=5, value=spp)
    ws4.cell(row=row, column=5).alignment = LEFT
    row += 1

style_body(ws4, 4, row - 1, len(ws4_headers))
auto_width(ws4)
ws4.column_dimensions['B'].width = 55
ws4.column_dimensions['C'].width = 35
ws4.column_dimensions['E'].width = 40

# ---- Sheet 5: 完整数据集 ----
ws5 = wb.create_sheet('完整数据集')

# 写表头
for i, col in enumerate(df.columns, 1):
    ws5.cell(row=1, column=i, value=col)
style_header(ws5, 1, len(df.columns))

# 写数据
for r_idx, (_, row_data) in enumerate(df.iterrows()):
    for c_idx, col in enumerate(df.columns):
        val = row_data[col]
        if pd.isna(val):
            val = ''
        ws5.cell(row=r_idx + 2, column=c_idx + 1, value=val)

style_body(ws5, 2, len(df) + 1, len(df.columns))
auto_width(ws5)
# 调宽备注列
ws5.column_dimensions[get_column_letter(list(df.columns).index('notes') + 1)].width = 50
ws5.column_dimensions[get_column_letter(list(df.columns).index('reference') + 1)].width = 45

# 冻结首行
ws5.freeze_panes = 'A2'

# ============================================================
# 保存
# ============================================================
os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)
print(f'[OK] Excel 文档已生成 → {OUT}')
# Also save to desktop (笔记本专用，本机跳过)
# os.makedirs(os.path.dirname(DESKTOP_OUT), exist_ok=True)
# wb.save(DESKTOP_OUT)
# print(f'[OK] 桌面副本 → {DESKTOP_OUT}')
print(f'  包含 5 个工作表: 数据总览 | 物种×城市矩阵 | 物种×功能区矩阵 | 参考文献清单 | 完整数据集')
